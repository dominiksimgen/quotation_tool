import openpyxl, os


class Read_input:
    def __init__(self):
        self.here = os.path.dirname(os.path.abspath(__file__))
        os.chdir(self.here)
        self.wb = openpyxl.load_workbook("input.xlsx")
        self.sheet1 = self.wb[self.wb.sheetnames[0]]
        self.requester_name = self.sheet1['B1'].value
        self.requester_contacts = self.sheet1['B2'].value
        self.stackable = self.sheet1['B3'].value
        self.dangerous_goods = self.sheet1['B4'].value
        self.li_ion = self.sheet1['B5'].value
        self.shipper_code = self.sheet1['B6'].value
        self.destination_code = self.sheet1['B7'].value
        self.material_dict = {}
        for i in range(10,16):
            if ((self.sheet1.cell(i,1).value) and (self.sheet1.cell(i,2).value) is not None):
                self.material_dict.update({self.sheet1.cell(i,1).value: self.sheet1.cell(i,2).value})