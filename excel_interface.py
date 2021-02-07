import openpyxl, os


class Read_excel_input:
    def __init__(self):
        self.here = os.path.dirname(os.path.abspath(__file__))
        os.chdir(self.here)
        self.wb = openpyxl.load_workbook("input.xlsx")
        self.sheet1 = self.wb[self.wb.sheetnames[0]]
        self.requester_name = self.sheet1['B1'].value
        self.requester_contacts = self.sheet1['B2'].value
        self.stackable = self.sheet1['B3'].value
        self.dangerous_goods = self.sheet1['B4'].value
        self.li_ion = self.sheet1['B5'].value
        self.shipper_code = str(self.sheet1['B6'].value)
        self.destination_code = str(self.sheet1['B7'].value)
        self.material_dict = {}
        for i in range(10,16):
            if ((self.sheet1.cell(i,1).value) and (self.sheet1.cell(i,2).value) is not None):
                self.material_dict.update({self.sheet1.cell(i,1).value: self.sheet1.cell(i,2).value})



class Form_Output:
    def __init__(self,form_template):
        self.here = os.path.dirname(os.path.abspath(__file__))
        os.chdir(self.here)
        self.wb = openpyxl.load_workbook(f"excel_files/{form_template}")
        self.sheet1 = self.wb[self.wb.sheetnames[0]]

    def fill_form(self, request):
        self.sheet1['B11'].value = str(request.reference_ID)[0:16]
        self.sheet1['H11'].value = request.requester_name
        self.sheet1['A18'].value = request.requester_name
        self.sheet1['H12'].value = request.requester_contacts
        self.sheet1['B50'].value = request.stackable
        self.sheet1['B52'].value = request.dangerous_goods
        self.sheet1['B56'].value = request.li_ion
        self.sheet1['H68'].value = request.ready_date

        #shipper information:
        self.sheet1['B78'].value = request.shipper_company_name
        self.sheet1['B79'].value = request.shipper_address['Street']
        self.sheet1['B80'].value = request.shipper_address['Postal Code / City']
        self.sheet1['B81'].value = request.shipper_address['Country']
        #destination information:
        self.sheet1['B59'].value = request.incoterm
        self.sheet1['D59'].value = request.airport
        self.sheet1['B84'].value = request.destination_company_name
        self.sheet1['B85'].value = request.destination_address['Street']
        self.sheet1['B86'].value = request.destination_address['Postal Code / City']
        self.sheet1['B87'].value = request.destination_address['Country']
        #list of shipment items:
        self.fill_material_list(request)
    
    def fill_material_list(self,request):
        self.material_row_counter = 28
        self.material_description_row_counter = self.material_row_counter + 1
        self.material_quantity_row_counter = self.material_row_counter + 2
        for material in request.material_object_array:
            self.sheet1.cell(self.material_row_counter,1).value = material.pallet_count
            self.sheet1.cell(self.material_row_counter,2).value = material.weight
            self.sheet1.cell(self.material_row_counter,3).value = material.fpc
            self.sheet1.cell(self.material_row_counter,5).value = material.packaging
            self.sheet1.cell(self.material_row_counter,7).value = material.pallet_length
            self.sheet1.cell(self.material_row_counter,9).value = material.pallet_width
            self.sheet1.cell(self.material_row_counter,11).value = material.pallet_height_airfreight
            self.sheet1.cell(self.material_description_row_counter, 3).value = material.description
            self.sheet1.cell(self.material_quantity_row_counter, 3).value = material.quantity
            self.material_row_counter += 3
            self.material_description_row_counter += 3
            self.material_quantity_row_counter += 3
        

    def save_output_to_excel(self, reference_ID):
        self.wb.save(f"output/{reference_ID}.xlsx")